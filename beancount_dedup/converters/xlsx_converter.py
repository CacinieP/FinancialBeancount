"""
XLSX 转换器 - 将 Excel 文件转换为 CSV
"""

from typing import List, Dict, Any, Optional
from pathlib import Path

from .base import BaseConverter, ConversionResult


class XLSXConverter(BaseConverter):
    """
    XLSX/XLS 到 CSV 转换器

    支持从 Excel 文件中提取数据并转换为 CSV 格式
    """

    @property
    def supported_extensions(self) -> List[str]:
        return ['.xlsx', '.xls']

    def convert(
        self,
        input_path: str,
        output_path: Optional[str] = None,
        sheet_name: Optional[str] = None,
        sheet_index: int = 0,
        **kwargs
    ) -> ConversionResult:
        """
        转换 Excel 文件为 CSV

        Args:
            input_path: Excel 文件路径
            output_path: 输出 CSV 文件路径
            sheet_name: 指定工作表名称（优先于 sheet_index）
            sheet_index: 工作表索引（默认 0，即第一个工作表）
            **kwargs: 额外参数
                - encoding: 输出编码（默认 utf-8-sig）
                - header_row: 表头行索引（默认 0）

        Returns:
            ConversionResult 转换结果
        """
        result = ConversionResult()

        # 验证输入文件
        if not Path(input_path).exists():
            result.errors.append(f"输入文件不存在: {input_path}")
            return result

        if not self.can_convert(input_path):
            result.errors.append(f"不支持的文件格式: {input_path}")
            return result

        # 生成输出路径
        if not output_path:
            output_path = self._generate_output_path(input_path)
        result.output_path = output_path

        # 提取数据
        rows = self._extract_excel_data(
            input_path,
            sheet_name,
            sheet_index,
            result,
            **kwargs
        )

        if not rows:
            result.errors.append("未能从 Excel 文件中提取到数据")
            return result

        # 写入 CSV
        try:
            encoding = kwargs.get('encoding', 'utf-8-sig')
            result.rows_converted = self._write_csv(rows, output_path, encoding)
            result.success = True
        except Exception as e:
            result.errors.append(f"写入 CSV 文件失败: {str(e)}")
            return result

        return result

    def _extract_excel_data(
        self,
        excel_path: str,
        sheet_name: Optional[str],
        sheet_index: int,
        result: ConversionResult,
        **kwargs
    ) -> List[Dict[str, Any]]:
        """
        从 Excel 文件中提取数据

        尝试多种方法按优先级：
        1. pandas - 最方便，支持多种格式
        2. openpyxl - 轻量级，仅支持 .xlsx
        3. xlrd - 支持 .xls（老版本）
        """
        # 方法 1: 尝试 pandas
        rows = self._try_pandas(
            excel_path,
            sheet_name,
            sheet_index,
            result,
            **kwargs
        )
        if rows:
            result.metadata['method'] = 'pandas'
            return rows

        # 方法 2: 尝试 openpyxl (仅 .xlsx)
        if excel_path.endswith('.xlsx'):
            rows = self._try_openpyxl(
                excel_path,
                sheet_name,
                sheet_index,
                result,
                **kwargs
            )
            if rows:
                result.metadata['method'] = 'openpyxl'
                return rows

        # 方法 3: 尝试 xlrd (仅 .xls)
        if excel_path.endswith('.xls'):
            rows = self._try_xlrd(
                excel_path,
                sheet_name,
                sheet_index,
                result,
                **kwargs
            )
            if rows:
                result.metadata['method'] = 'xlrd'
                return rows

        result.errors.append("所有 Excel 读取方法均失败")
        return []

    def _try_pandas(
        self,
        excel_path: str,
        sheet_name: Optional[str],
        sheet_index: int,
        result: ConversionResult,
        **kwargs
    ) -> Optional[List[Dict[str, Any]]]:
        """尝试使用 pandas 读取 Excel"""
        try:
            import pandas as pd
        except ImportError:
            result.warnings.append("pandas 未安装，跳过此方法")
            return None

        try:
            header_row = kwargs.get('header_row', 0)

            # 读取 Excel
            if sheet_name:
                df = pd.read_excel(
                    excel_path,
                    sheet_name=sheet_name,
                    header=header_row
                )
            else:
                df = pd.read_excel(
                    excel_path,
                    sheet_name=sheet_index,
                    header=header_row
                )

            # 转换为字典列表
            rows = []
            headers = self._normalize_headers(df.columns.tolist())

            for _, row in df.iterrows():
                row_dict = {}
                for header in headers:
                    value = row.get(header, '')
                    # 处理 NaN 值
                    if pd.isna(value):
                        value = ''
                    else:
                        value = str(value).strip()
                    row_dict[header] = value
                rows.append(row_dict)

            return rows

        except Exception as e:
            result.warnings.append(f"pandas 读取失败: {str(e)}")
            return None

    def _try_openpyxl(
        self,
        excel_path: str,
        sheet_name: Optional[str],
        sheet_index: int,
        result: ConversionResult,
        **kwargs
    ) -> Optional[List[Dict[str, Any]]]:
        """尝试使用 openpyxl 读取 Excel"""
        try:
            import openpyxl
        except ImportError:
            result.warnings.append("openpyxl 未安装，跳过此方法")
            return None

        try:
            header_row = kwargs.get('header_row', 0)

            workbook = openpyxl.load_workbook(excel_path, read_only=True)

            # 选择工作表
            if sheet_name:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.worksheets[sheet_index]

            rows = []
            headers = []

            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == header_row:
                    # 表头行
                    headers = self._normalize_headers([
                        str(cell) if cell is not None else ""
                        for cell in row
                    ])
                elif i > header_row:
                    # 数据行
                    row_dict = {}
                    for j, value in enumerate(row):
                        if j < len(headers):
                            header = headers[j]
                            row_dict[header] = str(value).strip() if value is not None else ""
                    rows.append(row_dict)

            workbook.close()
            return rows

        except Exception as e:
            result.warnings.append(f"openpyxl 读取失败: {str(e)}")
            return None

    def _try_xlrd(
        self,
        excel_path: str,
        sheet_name: Optional[str],
        sheet_index: int,
        result: ConversionResult,
        **kwargs
    ) -> Optional[List[Dict[str, Any]]]:
        """尝试使用 xlrd 读取旧版 Excel"""
        try:
            import xlrd
        except ImportError:
            result.warnings.append("xlrd 未安装，跳过此方法")
            return None

        try:
            header_row = kwargs.get('header_row', 0)

            workbook = xlrd.open_workbook(excel_path)

            # 选择工作表
            if sheet_name:
                sheet = workbook.sheet_by_name(sheet_name)
            else:
                sheet = workbook.sheet_by_index(sheet_index)

            rows = []
            headers = []

            for i in range(sheet.nrows):
                row = [str(sheet.cell_value(i, j)) for j in range(sheet.ncols)]

                if i == header_row:
                    # 表头行
                    headers = self._normalize_headers(row)
                elif i > header_row:
                    # 数据行
                    row_dict = {}
                    for j, value in enumerate(row):
                        if j < len(headers):
                            header = headers[j]
                            row_dict[header] = value.strip()
                    rows.append(row_dict)

            return rows

        except Exception as e:
            result.warnings.append(f"xlrd 读取失败: {str(e)}")
            return None

    def list_sheets(self, excel_path: str) -> List[str]:
        """
        列出 Excel 文件中的所有工作表名称

        Args:
            excel_path: Excel 文件路径

        Returns:
            工作表名称列表
        """
        if not Path(excel_path).exists():
            return []

        # 尝试使用 pandas
        try:
            import pandas as pd
            xl_file = pd.ExcelFile(excel_path)
            return xl_file.sheet_names
        except ImportError:
            pass

        # 尝试使用 openpyxl (仅 .xlsx)
        if excel_path.endswith('.xlsx'):
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(excel_path, read_only=True)
                sheets = workbook.sheetnames
                workbook.close()
                return sheets
            except ImportError:
                pass

        # 尝试使用 xlrd (仅 .xls)
        if excel_path.endswith('.xls'):
            try:
                import xlrd
                workbook = xlrd.open_workbook(excel_path)
                sheets = [sheet.name for sheet in workbook.sheets()]
                return sheets
            except ImportError:
                pass

        return []
